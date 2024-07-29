import csv
import collections
import openpyxl
from openpyxl.styles import PatternFill

def read_file(file, delimiter, primary_key_indices):
    data = collections.defaultdict(list)
    with open(file, 'r') as f:
        reader = csv.reader(f, delimiter=delimiter)
        header = next(reader)
        for row in reader:
            key = tuple(row[i] for i in primary_key_indices)
            data[key].append(row)
    return data, header

def read_mapping(mapping_file):
    mapping = {}
    with open(mapping_file, 'r') as f:
        reader = csv.reader(f)
        next(reader)  # Skip header
        for row in reader:
            file1_col, file2_col = row
            mapping[file1_col] = file2_col
    return mapping

def compare_files(file1, file2, primary_keys, delimiter, mapping_file):
    # Read column mapping
    column_mapping = read_mapping(mapping_file)

    # Read primary keys and full data from both files
    with open(file1, 'r') as f:
        reader = csv.reader(f, delimiter=delimiter)
        header1 = next(reader)
        primary_key_indices = [header1.index(key) for key in primary_keys]
        file1_data, header1 = read_file(file1, delimiter, primary_key_indices)
    
    with open(file2, 'r') as f:
        reader = csv.reader(f, delimiter=delimiter)
        header2 = next(reader)
        primary_key_indices2 = [header2.index(column_mapping[key]) for key in primary_keys]
        file2_data, header2 = read_file(file2, delimiter, primary_key_indices2)
    
    # Identify the columns to be processed in each chunk based on the mapping
    file1_columns = [header1.index(col) for col in header1 if col not in primary_keys]
    file2_columns = [header2.index(column_mapping[header1[col]]) for col in file1_columns]

    chunk_size = 10

    for start in range(0, len(file1_columns), chunk_size):
        end = start + chunk_size
        chunk_file1_columns = file1_columns[start:end]
        chunk_file2_columns = file2_columns[start:end]

        differences = []

        # Compare the data for the current chunk of columns
        for key in file1_data:
            if key not in file2_data:
                differences.append({
                    'type': 'missing_key',
                    'key': key,
                    'file': 'file2'
                })
            else:
                for row1 in file1_data[key]:
                    matched = False
                    for row2 in file2_data[key]:
                        if all(row1[i] == row2[j] for i, j in zip(chunk_file1_columns, chunk_file2_columns)):
                            matched = True
                            break
                    if not matched:
                        differences.append({
                            'type': 'row_mismatch',
                            'key': key,
                            'file1_row': [row1[i] for i in chunk_file1_columns],
                            'file2_rows': [[row2[j] for j in chunk_file2_columns] for row2 in file2_data[key]]
                        })

        for key in file2_data:
            if key not in file1_data:
                differences.append({
                    'type': 'missing_key',
                    'key': key,
                    'file': 'file1'
                })

        # Write the results to an Excel file
        #output_file = f"{output_file_prefix}_chunk_{start+1}_to_{end}.xlsx"
        outName = ', '.join(str(x) for x in chunk_file1_columns)
        outName = outName.replace(" ", "")
        output_file = f"Result_for_Col[{outName}].xlsx"
  
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = 'Differences'

        # Write the header row
        chunk_headers = [header1[i] for i in chunk_file1_columns]
        header_row = ['Type', 'Key', 'File', 'Description'] + chunk_headers
        sheet.append(header_row)

        # Write the differences
        for difference in differences:
            if difference['type'] == 'missing_key':
                row = [difference['type'], ', '.join(str(x) for x in difference['key']), difference['file'], f"Key {', '.join(str(x) for x in difference['key'])} not found in {difference['file']}"]
                sheet.append(row)
            elif difference['type'] == 'row_mismatch':
                row = [difference['type'], ', '.join(str(x) for x in difference['key']), 'both', f"Rows for key {', '.join(str(x) for x in difference['key'])} differ between files"]
                sheet.append(row)

                # Write the rows from each file once
                row1 = difference['file1_row']
                sheet.append(['', '', 'file1'] + row1)
                for row2 in difference['file2_rows']:
                    sheet.append(['', '', 'file2'] + row2)

                    # Highlight the differences and note the column names
                    column_names_with_diff = []
                    for j, (cell1, cell2) in enumerate(zip(row1, row2)):
                        if cell1 != cell2:
                            column_name = chunk_headers[j]
                            column_names_with_diff.append(column_name)
                            fill1 = PatternFill(start_color='c9e1f8', end_color='c9e1f8', fill_type='solid')
                            fill2 = PatternFill(start_color='f4cccc', end_color='f4cccc', fill_type='solid')
                            sheet.cell(row=sheet.max_row-1, column=4+j).fill = fill1
                            sheet.cell(row=sheet.max_row, column=4+j).fill = fill2

                    # Add an additional row for column names with differences
                    description = f"Columns with differences: {', '.join(column_names_with_diff)}"
                    sheet.append(['', '', '', description])

        # Save the workbook for the current chunk
        workbook.save(output_file)
        print(f"Results for columns {start+1} to {end} written to {output_file}")

# Get the file names, primary keys, and delimiter from the user
# file1 = input("Enter the first file name: ")
# file2 = input("Enter the second file name: ")
# primary_keys = input("Enter the primary key column names (comma-separated): ").split(',')
# delimiter = input("Enter the delimiter (e.g. , or ;): ")
file1 = "InputFiles\sample1.csv"
file2 = "InputFiles\sample2.csv"
primary_keys = ["OrderID", "Region"]
delimiter = ","
mapping = 'mapping.csv'

# Compare the files
compare_files(file1, file2, primary_keys, delimiter, mapping)